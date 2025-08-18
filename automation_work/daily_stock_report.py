import yfinance as yf  # type: ignore
import pandas as pd  # type: ignore
import numpy as np
from datetime import datetime, timedelta, date
from pathlib import Path
from openpyxl.styles import PatternFill # type: ignore
from openpyxl.formatting.rule import CellIsRule, FormulaRule # type: ignore
from openpyxl.utils import get_column_letter # type: ignore


# tickers to report
tickers = [
    "SIE.DE", "ALV.DE", "SAP.DE", "DTE.DE", "BAS.DE", "BAYN.DE", "DBK.DE",
    "IFX.DE", "BMW.DE", "MBG.DE", "AIR.PA", "DPW.DE", "CBK.DE", "RWE.DE", "VOW3.DE"
]

# reference date: Friday 15.8. of the current year (fallback to previous year if in future)
today = date.today()
ref_day = 15
ref_month = 8
ref_year = today.year
ref_date = date(ref_year, ref_month, ref_day)
if ref_date > today:
    ref_date = date(ref_year - 1, ref_month, ref_day)

# build list of calendar dates from reference -> today (inclusive)
calendar_dates = [ref_date + timedelta(days=i) for i in range((today - ref_date).days + 1)]

# skip 16.8 and 17.8
skip_month = 8
skip_days = {16, 17}
filtered_dates = [d for d in calendar_dates if not (d.month == skip_month and d.day in skip_days)]

# column names:
# first column is reference price (Ref_DD.MM.YYYY_Price),
# then for each subsequent date we create two columns: <DD.MM.YYYY>_Price and <DD.MM.YYYY>_Diff
date_fmt_header = "%d.%m.%Y"
col_names = []
col_names.append(f"Ref_{ref_date.strftime(date_fmt_header)}_Price")
for d in filtered_dates[1:]:
    label = d.strftime(date_fmt_header)
    col_names.append(f"{label}_Price")
    col_names.append(f"{label}_Diff")

MISSING = "MISSING VALUE"
rows = {}
names = {}  # ticker -> stock name cache

def fmt_val(v):
    """Return rounded float to 2 decimals or MISSING."""
    if v is None:
        return MISSING
    try:
        f = float(v)
    except Exception:
        return MISSING
    if np.isnan(f):
        return MISSING
    return round(f, 2)

# fetch data per ticker
for ticker in tickers:
    values = []
    try:
        stock = yf.Ticker(ticker)
        # try to get a readable name; fallback to ticker
        try:
            info = stock.info
            short_name = info.get("shortName") or info.get("longName") or ticker
        except Exception:
            short_name = ticker
        names[ticker] = short_name

        # request history covering ref_date .. today (end is exclusive so add one day)
        hist = stock.history(start=ref_date.strftime("%Y-%m-%d"),
                             end=(today + timedelta(days=1)).strftime("%Y-%m-%d"))
        # create mapping date -> close (as float)
        close_map = {}
        if "Close" in hist and not hist["Close"].empty:
            for ts, val in hist["Close"].items():
                try:
                    close_map[ts.date()] = float(val)
                except Exception:
                    pass

        # iterate dates in order and build price + diff columns
        prev_raw = None
        # reference price
        raw_ref = close_map.get(ref_date, None)
        values.append(fmt_val(raw_ref))
        prev_raw = raw_ref  # prev for next day comparison

        for d in filtered_dates[1:]:
            raw = close_map.get(d, None)
            # price column
            values.append(fmt_val(raw))
            # diff column: current - prev (both raw numeric required)
            if (prev_raw is not None) and (raw is not None):
                diff = round(raw - prev_raw, 2)
                values.append(diff)
            else:
                values.append(MISSING)
            prev_raw = raw

        # if nothing at all present, warn
        if all(v == MISSING for v in values):
            print(f"Warning: no price data for {ticker} from {ref_date} to {today} (may be delisted).")
    except Exception as e:
        print(f"Error fetching {ticker}: {e}")
        values = [MISSING] * len(col_names)
        names[ticker] = ticker
    rows[ticker] = values

# build DataFrame
df = pd.DataFrame.from_dict(rows, orient="index", columns=col_names)
df.index.name = "Ticker"

# add Stock Name column and move it to first column
df.insert(0, "Stock Name", df.index.map(names))
df.reset_index(inplace=True)  # keep Ticker as a column

# ensure missing values are the requested string
df = df.replace({np.nan: MISSING})
df = df.fillna(MISSING)

# Save to requested folder with conditional formatting for > 1 EUR diffs
output_dir = Path(r"C:\Users\I070494\Downloads\EU_Akcie_AI_Ceny")
output_dir.mkdir(parents=True, exist_ok=True)
# filename with date in DD.MM.YYYY format
filename = f"stock_report_{today.strftime('%d.%m.%Y')}.xlsx"
out_path = output_dir / filename

with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
    df.to_excel(writer, index=False, sheet_name="Report")
    ws = writer.sheets["Report"]

    # find Diff columns
    yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    diff_cols = [(idx + 1, col_name) for idx, col_name in enumerate(df.columns) if col_name.endswith("_Diff")]
    last_row = len(df) + 1  # header row + data rows

    # apply formula-based conditional formatting that ignores non-numeric "MISSING VALUE"
    for col_idx, col_name in diff_cols:
        col_letter = get_column_letter(col_idx)
        cell_range = f"{col_letter}2:{col_letter}{last_row}"
        formula = f"=AND(ISNUMBER({col_letter}2), ABS({col_letter}2)>=1)"
        rule = FormulaRule(formula=[formula], fill=yellow)
        ws.conditional_formatting.add(cell_range, rule)

    # highlight Stock Name cell for rows where any diff numeric abs >= 1
    if diff_cols and "Stock Name" in df.columns:
        diff_letters = [get_column_letter(col_idx) for col_idx, _ in diff_cols]
        conditions = [f"AND(ISNUMBER(${col}2),ABS(${col}2)>=1)" for col in diff_letters]
        formula = f"=OR({','.join(conditions)})"
        name_col_idx = df.columns.get_loc("Stock Name") + 1
        name_col_letter = get_column_letter(name_col_idx)
        name_range = f"{name_col_letter}2:{name_col_letter}{last_row}"
        fr = FormulaRule(formula=[formula], fill=yellow)
        ws.conditional_formatting.add(name_range, fr)

        # Autosize columns to fit content
        for col_cells in ws.columns:
            try:
                max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col_cells)
            except Exception:
                max_length = 0
            col_letter = get_column_letter(col_cells[0].column)
            adjusted_width = min(max_length + 2, 60)  # add padding, cap width
            ws.column_dimensions[col_letter].width = adjusted_width

print(f"Saved report to: {out_path}")
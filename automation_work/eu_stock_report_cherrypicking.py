"""
Simplified EU stock reporter (cherry-picking).

Creates/updates an Excel workbook at:
  C:/Users/I070494/Downloads/AI_EU_StockPrices_Report/EU_StockPrices_Report_CherryPicking.xlsx

Each run:
 - selects 10-15 stable EUR stocks from a candidate list (stability = max +/-10% dev from 60-day mean)
 - creates a new sheet named dd.mm.YYYY (if exists appends HHMMSS)
 - writes 5 columns: Ticker (hyperlink), Stock Name, Previous Day Closing price, Actual price, Price Diff
 - highlights Actual price & Price Diff in green font when abs(diff) >= 1.0 EUR
 - autosizes columns and removes default Sheet1
"""
from pathlib import Path
from datetime import datetime
import math

import yfinance as yf # type: ignore
import pandas as pd # type: ignore
from openpyxl import load_workbook # type: ignore
from openpyxl.styles import Font # type: ignore
from openpyxl.formatting.rule import FormulaRule # type: ignore
from openpyxl.utils import get_column_letter # type: ignore

# Configuration
OUT_DIR = Path(r"C:\Users\I070494\Downloads\AI_EU_StockPrices_Report")
OUT_DIR.mkdir(parents=True, exist_ok=True)
MASTER_FILE = OUT_DIR / "EU_StockPrices_Report_CherryPicking.xlsx"

CANDIDATES = [
    "SAP.DE", "SIE.DE", "BMW.DE", "MBG.DE", "ALV.DE",
    "BAS.DE", "BAYN.DE", "VOW3.DE", "RWE.DE", "DTE.DE",
    "AIR.PA", "ASML.AS", "MC.PA", "OR.PA", "BNP.PA"
]

LOOKBACK_DAYS = 60
MAX_REL_DEV = 0.10
HIGHLIGHT_THRESHOLD = 1.0
MISSING = "MISSING VALUE"

# --- Helper functions -------------------------------------------------------
def fetch_close_series(ticker: str, days: int):
    """Return pandas Series of Close prices (may be empty)."""
    try:
        tk = yf.Ticker(ticker)
        hist = tk.history(period=f"{days + 10}d", auto_adjust=False)
        if "Close" in hist:
            return hist["Close"].dropna()
    except Exception:
        pass
    return pd.Series(dtype=float)

def is_stable(ticker: str) -> bool:
    """Check stability: max relative deviation from mean over LOOKBACK_DAYS <= MAX_REL_DEV."""
    s = fetch_close_series(ticker, LOOKBACK_DAYS)
    if s.empty:
        return False
    s = s.tail(LOOKBACK_DAYS)
    mean = s.mean()
    if mean == 0 or math.isnan(mean):
        return False
    rel = (s - mean).abs() / mean
    return rel.max() <= MAX_REL_DEV

def pick_tickers(candidates, min_n=10, max_n=15):
    """Return up to max_n stable tickers, fill up to min_n if not enough pass."""
    picked = []
    for t in candidates:
        try:
            if is_stable(t):
                picked.append(t)
        except Exception:
            continue
        if len(picked) >= max_n:
            break
    if len(picked) < min_n:
        for t in candidates:
            if t not in picked:
                picked.append(t)
            if len(picked) >= min_n:
                break
    return picked[:max_n]

def fetch_prev_and_actual(ticker: str):
    """Return (prev_close or None, actual_close or None) using last available closes."""
    s = fetch_close_series(ticker, 5)
    if s.empty:
        return None, None
    s = s.tail(3)
    if len(s) >= 2:
        return float(s.iloc[-2]), float(s.iloc[-1])
    if len(s) == 1:
        return None, float(s.iloc[-1])
    return None, None

def fetch_name(ticker: str) -> str:
    """Try to fetch stock readable name; fallback to ticker symbol."""
    try:
        info = yf.Ticker(ticker).info
        return info.get("shortName") or info.get("longName") or ticker
    except Exception:
        return ticker

def safe_round(v):
    """Round numeric to 2 decimals or return MISSING for invalid values."""
    if v is None:
        return MISSING
    try:
        if isinstance(v, float) and math.isnan(v):
            return MISSING
        return round(float(v), 2)
    except Exception:
        return MISSING

def autosize_ws(ws):
    """Autosize columns using a simple heuristic."""
    for col_cells in ws.columns:
        max_len = 0
        for cell in col_cells:
            try:
                val = cell.value
                l = 0 if val is None else len(str(val))
                if l > max_len:
                    max_len = l
            except Exception:
                pass
        col_letter = get_column_letter(col_cells[0].column)
        ws.column_dimensions[col_letter].width = min(max_len + 3, 60)

# --- Build DataFrame for the run -------------------------------------------
def build_run_dataframe(tickers):
    """Construct DataFrame with required columns for the provided tickers."""
    rows = []
    for t in tickers:
        prev, actual = None, None
        try:
            prev, actual = fetch_prev_and_actual(t)
        except Exception:
            prev, actual = None, None
        prev_val = safe_round(prev)
        actual_val = safe_round(actual)
        if isinstance(prev_val, (int, float)) and isinstance(actual_val, (int, float)):
            diff = round(actual_val - prev_val, 2)
        else:
            diff = MISSING
        name = fetch_name(t)
        rows.append({
            "Ticker": t,
            "Stock Name": name,
            "Previous Day Closing price": prev_val,
            "Actual price": actual_val,
            "Price Diff": diff
        })
    return pd.DataFrame(rows, columns=[
        "Ticker", "Stock Name", "Previous Day Closing price", "Actual price", "Price Diff"
    ])

# --- Write sheet to Excel --------------------------------------------------
def write_report_sheet(df: pd.DataFrame):
    """
    Add new sheet named dd.mm.YYYY (unique) to MASTER_FILE and post-process:
    - timestamp in A1
    - hyperlink tickers to Yahoo Finance
    - conditional formatting: green font if abs(diff) >= threshold
    - autosize columns and remove default Sheet1
    """
    now = datetime.now()
    sheet_name = now.strftime("%d.%m.%Y")

    # decide writer mode and ensure unique sheet name
    if MASTER_FILE.exists():
        wb = load_workbook(MASTER_FILE)
        if sheet_name in wb.sheetnames:
            sheet_name = now.strftime("%d.%m.%Y_%H%M%S")
        wb.close()
        mode = "a"
    else:
        mode = "w"

    writer_kwargs = {"engine": "openpyxl", "mode": mode}
    if mode == "a":
        writer_kwargs["if_sheet_exists"] = "new"

    # write DataFrame so header appears on row 2 (startrow=1)
    with pd.ExcelWriter(MASTER_FILE, **writer_kwargs) as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=1)

    # open workbook for formatting
    wb = load_workbook(MASTER_FILE)
    # remove default Sheet1 if present
    if "Sheet1" in wb.sheetnames:
        try:
            wb.remove(wb["Sheet1"])
        except Exception:
            pass

    ws = wb[sheet_name]
    ws["A1"] = f"Run at: {now.strftime('%d.%m.%Y %H:%M:%S')}"
    header_row = 2
    data_start = header_row + 1
    last_row = ws.max_row

    headers = [cell.value for cell in ws[header_row]]
    # ticker column index (1-based), fallback to 1
    try:
        ticker_idx = headers.index("Ticker") + 1
    except ValueError:
        ticker_idx = 1

    # add Yahoo hyperlinks to ticker cells
    for r in range(data_start, last_row + 1):
        cell = ws.cell(row=r, column=ticker_idx)
        if cell.value:
            try:
                url = f"https://finance.yahoo.com/quote/{cell.value}"
                cell.hyperlink = url
                cell.font = Font(color="0000FF", underline="single")
            except Exception:
                pass

    # find Actual price and Price Diff columns
    try:
        act_idx = headers.index("Actual price") + 1
        diff_idx = headers.index("Price Diff") + 1
    except ValueError:
        wb.save(MASTER_FILE)
        wb.close()
        return

    act_col = get_column_letter(act_idx)
    diff_col = get_column_letter(diff_idx)

    # conditional formatting: green font where abs(diff) >= HIGHLIGHT_THRESHOLD and numeric
    green = Font(color="008000")
    formula = f"AND(ISNUMBER(${diff_col}{data_start}),ABS(${diff_col}{data_start})>={HIGHLIGHT_THRESHOLD})"
    rule = FormulaRule(formula=[formula], font=green)

    ws.conditional_formatting.add(f"{diff_col}{data_start}:{diff_col}{last_row}", rule)
    ws.conditional_formatting.add(f"{act_col}{data_start}:{act_col}{last_row}", rule)

    autosize_ws(ws)
    wb.save(MASTER_FILE)
    wb.close()

# --- Main ------------------------------------------------------------------
def main():
    tickers = pick_tickers(CANDIDATES, min_n=10, max_n=15)
    if not tickers:
        print("No tickers selected after filtering. Exiting.")
        return
    df = build_run_dataframe(tickers)
    write_report_sheet(df)
    print(f"Report written to: {MASTER_FILE}")

if __name__ == "__main__":
    main()
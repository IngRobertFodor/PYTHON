"""
Simple script to fetch EUR-denominated prices for selected stable European stocks,
create a new Excel sheet named dd.mm.YYYY for each run and store:
  1) Ticker
  2) Stock Name
  3) Previous Day Closing price
  4) Actual price (numeric)  -- run timestamp is written above the table
  5) Price Diff (Actual - Previous)
Highlight Actual price and Price Diff in green font when abs(diff) >= 1 EUR.
Autosize columns. Create target folder if missing and remove default Sheet1.
Clear comments describe each section.
"""

# --------------------------
# Imports and configuration
# --------------------------
import sys
from pathlib import Path
from datetime import datetime, date
import math

import yfinance as yf  # type: ignore # pip install yfinance
import pandas as pd    # type: ignore # pip install pandas openpyxl
from openpyxl import load_workbook # type: ignore
from openpyxl.styles import Font # type: ignore
from openpyxl.formatting.rule import FormulaRule # type: ignore
from openpyxl.utils import get_column_letter # type: ignore

# Output folder and master file
OUT_DIR = Path(r"C:\Users\I070494\Downloads\AI_EU_StockPrices_Report")
OUT_DIR.mkdir(parents=True, exist_ok=True)
MASTER_FILE = OUT_DIR / "EU_StockPrices_Report.xlsx"

# Choose 10-15 stable European stocks (EUR-listed or EUR-traded tickers on major exchanges)
TICKERS = [
    "SAP.DE",   # SAP SE (Germany, EUR)
    "SIE.DE",   # Siemens AG (Germany, EUR)
    "BMW.DE",   # BMW AG (Germany, EUR)
    "MBG.DE",   # Mercedes-Benz Group AG (Germany, EUR)
    "ALV.DE",   # Allianz SE (Germany, EUR)
    "BAS.DE",   # BASF SE (Germany, EUR)
    "BAYN.DE",  # Bayer AG (Germany, EUR)
    "VOW3.DE",  # Volkswagen AG (Germany, EUR)
    "RWE.DE",   # RWE AG (Germany, EUR)
    "DTE.DE",   # Deutsche Telekom AG (Germany, EUR)
    "AIR.PA",   # Airbus SE (France, EUR)
    "AIR.PA",   # Airbus repeated removed by code later if duplicate
]

# Ensure unique tickers
TICKERS = list(dict.fromkeys(TICKERS))

# Threshold for highlighting (EUR)
HIGHLIGHT_THRESHOLD = 1.0

# --------------------------
# Helper functions
# --------------------------
def fetch_prev_and_actual_close(ticker):
    """
    Fetch last available closes for given ticker.
    Returns tuple (prev_close_or_None, actual_close_or_None).
    Uses yfinance history for robustness.
    """
    try:
        tk = yf.Ticker(ticker)
        # get last 3 trading days to be safe
        hist = tk.history(period="3d", auto_adjust=False)
        closes = hist["Close"].dropna()
        if len(closes) >= 2:
            prev = float(closes.iloc[-2])
            actual = float(closes.iloc[-1])
            return prev, actual
        elif len(closes) == 1:
            # only one available - treat prev as None
            actual = float(closes.iloc[-1])
            return None, actual
        else:
            return None, None
    except Exception:
        return None, None

def fetch_stock_name(ticker):
    """Try to fetch a human readable stock name; fallback to ticker."""
    try:
        info = yf.Ticker(ticker).info
        return info.get("shortName") or info.get("longName") or ticker
    except Exception:
        return ticker

def safe_round(v):
    """Round numeric to 2 decimals; return 'MISSING VALUE' for None/NaN."""
    if v is None:
        return "MISSING VALUE"
    try:
        if isinstance(v, float) and math.isnan(v):
            return "MISSING VALUE"
        return round(float(v), 2)
    except Exception:
        return "MISSING VALUE"

def autosize_workbook_columns(ws):
    """Autosize all columns in worksheet based on cell contents."""
    for col_cells in ws.columns:
        max_length = 0
        for cell in col_cells:
            try:
                val = cell.value
                if val is None:
                    length = 0
                else:
                    length = len(str(val))
                if length > max_length:
                    max_length = length
            except Exception:
                pass
        col_letter = get_column_letter(col_cells[0].column)
        # add small padding, cap to avoid extremely wide columns
        ws.column_dimensions[col_letter].width = min(max_length + 3, 60)

# --------------------------
# Build DataFrame for this run
# --------------------------
def build_run_dataframe(tickers):
    """
    Construct DataFrame with columns:
    ['Ticker','Stock Name','Previous Day Closing price','Actual price','Price Diff']
    """
    rows = []
    for t in tickers:
        prev, actual = fetch_prev_and_actual_close(t)
        prev_val = safe_round(prev)
        actual_val = safe_round(actual)
        # compute diff if both numeric
        if isinstance(prev_val, (int, float)) and isinstance(actual_val, (int, float)):
            diff = round(actual_val - prev_val, 2)
        else:
            diff = "MISSING VALUE"
        name = fetch_stock_name(t)
        rows.append({
            "Ticker": t,
            "Stock Name": name,
            "Previous Day Closing price": prev_val,
            "Actual price": actual_val,
            "Price Diff": diff,
        })
    df = pd.DataFrame(rows, columns=[
        "Ticker", "Stock Name", "Previous Day Closing price", "Actual price", "Price Diff"
    ])
    return df

# --------------------------
# Excel write logic
# --------------------------
def write_report_sheet(df):
    """
    Create or open MASTER_FILE, add new sheet named dd.mm.YYYY (if exists append time),
    write run timestamp above data, delete default Sheet1 if present,
    apply conditional formatting (green font) to Actual price and Price Diff when abs(diff) >= threshold,
    autosize columns.
    """
    now = datetime.now()
    sheet_name = now.strftime("%d.%m.%Y")

    # If master exists, append; otherwise create new workbook
    if MASTER_FILE.exists():
        wb = load_workbook(MASTER_FILE)
        if sheet_name in wb.sheetnames:
            # avoid collision: append time to keep unique sheet names if same-day rerun
            sheet_name = now.strftime("%d.%m.%Y_%H%M%S")
        wb.close()
        mode = "a"
    else:
        mode = "w"

    # prepare ExcelWriter kwargs depending on mode to avoid pandas ValueError
    writer_kwargs = {"engine": "openpyxl", "mode": mode}
    if mode == "a":
        # only valid when appending to existing file
        writer_kwargs["if_sheet_exists"] = "new"

    # write df to new sheet (start at row 2 so we can place timestamp on row 1)
    with pd.ExcelWriter(MASTER_FILE, **writer_kwargs) as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=1)

    # open workbook to post-process: remove default Sheet1 and apply formatting
    wb = load_workbook(MASTER_FILE)
    # delete default Sheet1 if present and it is empty (common on first create)
    if "Sheet1" in wb.sheetnames:
        try:
            std = wb["Sheet1"]
            # remove even if not strictly empty to follow user's request
            wb.remove(std)
        except Exception:
            pass

    ws = wb[sheet_name]

    # Write run timestamp in A1
    timestamp_text = f"Run at: {now.strftime('%d.%m.%Y %H:%M:%S')}"
    ws["A1"] = timestamp_text

    # Determine column letters for 'Actual price' and 'Price Diff'
    headers = [cell.value for cell in ws[2]]  # header is in row 2 because we started at row=1
    # find indices (1-based)
    try:
        act_idx = headers.index("Actual price") + 1
        diff_idx = headers.index("Price Diff") + 1
    except ValueError:
        # header mismatch - abort safely
        wb.save(MASTER_FILE)
        wb.close()
        return

    act_col = get_column_letter(act_idx)
    diff_col = get_column_letter(diff_idx)

    # Apply conditional formatting using formula that checks numeric and abs >= threshold.
    green_font = Font(color="008000")  # green

    last_row = ws.max_row
    # Range references (exclude header row: start at 3 because header row is 2 and data starts at 3)
    act_range = f"{act_col}3:{act_col}{last_row}"
    diff_range = f"{diff_col}3:{diff_col}{last_row}"

    # Formula rule for absolute diff >= HIGHLIGHT_THRESHOLD and numeric
    formula = f"AND(ISNUMBER(${diff_col}3),ABS(${diff_col}3)>={HIGHLIGHT_THRESHOLD})"
    rule = FormulaRule(formula=[formula], font=green_font)

    # Apply rule to both diff column and actual price column (actual uses same formula referencing diff cell)
    ws.conditional_formatting.add(diff_range, rule)
    formula_act = f"AND(ISNUMBER(${diff_col}3),ABS(${diff_col}3)>={HIGHLIGHT_THRESHOLD})"
    rule_act = FormulaRule(formula=[formula_act], font=green_font)
    ws.conditional_formatting.add(act_range, rule_act)

    # Autosize columns
    autosize_workbook_columns(ws)

    # Save and close workbook
    wb.save(MASTER_FILE)
    wb.close()

# --------------------------
# Main
# --------------------------
def main():
    # Build DataFrame for current run
    df = build_run_dataframe(TICKERS)
    # Write to Excel (new sheet)
    write_report_sheet(df)

# For standalone run (not imported as module)
if __name__ == "__main__":
    main()